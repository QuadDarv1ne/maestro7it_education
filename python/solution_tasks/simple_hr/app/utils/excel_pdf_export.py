"""
Excel и PDF экспорт для отчетов
"""
from io import BytesIO
from datetime import datetime
from flask import make_response

def create_excel_response(workbook, filename):
    """Создать HTTP ответ с Excel файлом"""
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response

def create_pdf_response(pdf_content, filename):
    """Создать HTTP ответ с PDF файлом"""
    response = make_response(pdf_content)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response

class ExcelExporter:
    """Экспорт данных в Excel"""
    
    @staticmethod
    def export_employees(employees):
        """Экспорт списка сотрудников в Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Сотрудники"
            
            # Заголовок
            headers = ['№', 'ФИО', 'Email', 'Табельный номер', 'Подразделение', 
                      'Должность', 'Дата приема', 'Статус']
            
            # Стиль заголовка
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Данные
            for idx, emp in enumerate(employees, 1):
                ws.cell(row=idx+1, column=1, value=idx)
                ws.cell(row=idx+1, column=2, value=emp.full_name)
                ws.cell(row=idx+1, column=3, value=emp.email)
                ws.cell(row=idx+1, column=4, value=emp.employee_id)
                ws.cell(row=idx+1, column=5, value=emp.department.name if emp.department else '')
                ws.cell(row=idx+1, column=6, value=emp.position.title if emp.position else '')
                ws.cell(row=idx+1, column=7, value=emp.hire_date.strftime('%d.%m.%Y') if emp.hire_date else '')
                ws.cell(row=idx+1, column=8, value='Активен' if emp.status == 'active' else 'Уволен')
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filename = f'employees_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            return create_excel_response(wb, filename)
            
        except ImportError:
            raise Exception("openpyxl не установлен. Установите: pip install openpyxl")
    
    @staticmethod
    def export_vacations(vacations):
        """Экспорт отпусков в Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Отпуска"
            
            headers = ['№', 'Сотрудник', 'Тип', 'Дата начала', 'Дата окончания', 
                      'Длительность (дней)', 'Статус']
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            vacation_types = {
                'paid': 'Оплачиваемый',
                'unpaid': 'Неоплачиваемый',
                'sick': 'Больничный'
            }
            
            status_types = {
                'pending': 'Ожидает',
                'approved': 'Одобрен',
                'rejected': 'Отклонен'
            }
            
            for idx, vac in enumerate(vacations, 1):
                duration = (vac.end_date - vac.start_date).days + 1 if vac.end_date and vac.start_date else 0
                
                ws.cell(row=idx+1, column=1, value=idx)
                ws.cell(row=idx+1, column=2, value=vac.employee.full_name if vac.employee else '')
                ws.cell(row=idx+1, column=3, value=vacation_types.get(vac.type, vac.type))
                ws.cell(row=idx+1, column=4, value=vac.start_date.strftime('%d.%m.%Y') if vac.start_date else '')
                ws.cell(row=idx+1, column=5, value=vac.end_date.strftime('%d.%m.%Y') if vac.end_date else '')
                ws.cell(row=idx+1, column=6, value=duration)
                ws.cell(row=idx+1, column=7, value=status_types.get(vac.status, vac.status))
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filename = f'vacations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            return create_excel_response(wb, filename)
            
        except ImportError:
            raise Exception("openpyxl не установлен")


class PDFExporter:
    """Экспорт данных в PDF"""
    
    @staticmethod
    def export_employees_report(employees, title="Отчет по сотрудникам"):
        """Экспорт отчета по сотрудникам в PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            buffer = BytesIO()
            
            # Создаем документ
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
            elements = []
            
            # Стили
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#366092'),
                spaceAfter=30,
                alignment=1  # Центрирование
            )
            
            # Заголовок
            elements.append(Paragraph(title, title_style))
            elements.append(Paragraph(f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            # Таблица данных
            data = [['№', 'ФИО', 'Email', 'Таб. №', 'Подразделение', 'Должность', 'Дата приема', 'Статус']]
            
            for idx, emp in enumerate(employees, 1):
                data.append([
                    str(idx),
                    emp.full_name[:30],
                    emp.email[:25],
                    emp.employee_id,
                    (emp.department.name[:20] if emp.department else '')[:20],
                    (emp.position.title[:20] if emp.position else '')[:20],
                    emp.hire_date.strftime('%d.%m.%Y') if emp.hire_date else '',
                    'Активен' if emp.status == 'active' else 'Уволен'
                ])
            
            # Создаем таблицу
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            elements.append(table)
            
            # Сводка
            elements.append(Spacer(1, 0.3*inch))
            summary = f"Всего сотрудников: {len(employees)}"
            elements.append(Paragraph(summary, styles['Normal']))
            
            doc.build(elements)
            
            pdf_content = buffer.getvalue()
            buffer.close()
            
            filename = f'employees_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            return create_pdf_response(pdf_content, filename)
            
        except ImportError:
            raise Exception("reportlab не установлен. Установите: pip install reportlab")
